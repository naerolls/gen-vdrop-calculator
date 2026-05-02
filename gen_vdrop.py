#!/usr/bin/env python3
"""
gen_vdrop.py
============
Generator Voltage Drop Calculator
Salient-Pole Newton-Raphson Solution + Excel Report Export

Covers:
  • Simplified cylindrical-rotor formula  (quick reference)
  • Full salient-pole NR solution         (Xd + Xq + Ra)
  • Step load study                       (no pre-existing load)
  • Motor starting study                  (pre-existing base load + motor inrush)
  • Unit conversions                      (HP, %, Ω, kVA, V, A)
  • Excel report export                   (openpyxl)

Usage
-----
  Interactive mode (prompted inputs):
      python gen_vdrop.py

  Step load study, CLI inputs:
      python gen_vdrop.py --mode step \\
          --gen-kva 8125 --gen-v 4160 \\
          --xd 1.913 --xdp 0.222 --xdpp 0.169 \\
          --xq 0.871 --xqp 0.871 --xqpp 0.222 --r1 0.0069 \\
          --load-kw 6500 --load-pf 0.80

  Motor starting study, CLI inputs:
      python gen_vdrop.py --mode motor \\
          --gen-kva 8125 --gen-v 4160 \\
          --xd 1.913 --xdp 0.222 --xdpp 0.169 \\
          --xq 0.871 --xqp 0.871 --xqpp 0.222 --r1 0.0069 \\
          --base-kva 720 --base-pf 0.85 \\
          --motor-kva 7164 --motor-pf 0.15

  Motor starting study, HP + FLA inputs (unit conversion built in):
      python gen_vdrop.py --mode motor \\
          --gen-kva 8125 --gen-v 4160 \\
          --xdpp 0.169 --xqpp 0.222 --xdp 0.222 --xqp 0.871 \\
          --xd 1.913 --xq 0.871 --r1 0.0069 \\
          --base-kva 720 --base-pf 0.85 \\
          --hp 1250 --fla 156 --v-motor 4160 --inrush-mult 6.5 --motor-pf 0.15

  Export Excel report:
      python gen_vdrop.py [<any inputs>] --export report.xlsx

  Standalone unit conversions:
      python gen_vdrop.py --convert hp 1250
      python gen_vdrop.py --convert pct 16.9
      python gen_vdrop.py --convert starting-kva 156 6.5 4160
      python gen_vdrop.py --convert ohms 0.071 8125 4160

Dependencies
------------
  Standard library only for calculations.
  openpyxl required for Excel export:  pip install openpyxl
"""

from __future__ import annotations
import argparse
import math
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

# ─────────────────────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────────────────────
HP_TO_KW = 0.74570      # 1 mechanical HP
SQRT3    = math.sqrt(3)
LINE     = '─' * 78


# ─────────────────────────────────────────────────────────────────────────────
# Unit Conversions
# ─────────────────────────────────────────────────────────────────────────────
class Units:
    """All unit conversion helpers, grouped for easy reference."""

    @staticmethod
    def hp_to_kw(hp: float) -> float:
        return hp * HP_TO_KW

    @staticmethod
    def kw_to_hp(kw: float) -> float:
        return kw / HP_TO_KW

    @staticmethod
    def pct_to_pu(pct: float) -> float:
        """Convert reactance from % at kVA rating to per-unit."""
        return pct / 100.0

    @staticmethod
    def pu_to_pct(pu: float) -> float:
        return pu * 100.0

    @staticmethod
    def base_impedance(kva_rated: float, voltage_ll_v: float) -> float:
        """Z_base = V_LL² / (kVA × 1000)  [Ohms]"""
        return (voltage_ll_v ** 2) / (kva_rated * 1000.0)

    @staticmethod
    def ohms_to_pu(z_ohms: float, kva_rated: float, voltage_ll_v: float) -> float:
        return z_ohms / Units.base_impedance(kva_rated, voltage_ll_v)

    @staticmethod
    def pu_to_ohms(z_pu: float, kva_rated: float, voltage_ll_v: float) -> float:
        return z_pu * Units.base_impedance(kva_rated, voltage_ll_v)

    @staticmethod
    def vll_to_vln(v_ll: float) -> float:
        return v_ll / SQRT3

    @staticmethod
    def vln_to_vll(v_ln: float) -> float:
        return v_ln * SQRT3

    @staticmethod
    def motor_starting_kva_from_fla(fla_amps: float,
                                     inrush_mult: float,
                                     voltage_ll_v: float) -> float:
        """S_start = √3 × V_LL × (FLA × inrush_mult) / 1000  [kVA]"""
        return SQRT3 * voltage_ll_v * fla_amps * inrush_mult / 1000.0

    @staticmethod
    def motor_starting_kva_from_hp(hp: float,
                                    efficiency: float,
                                    rated_pf: float,
                                    start_mult: float) -> float:
        """Approximate starting kVA from HP (less accurate — prefer FLA method)."""
        rated_kw  = Units.hp_to_kw(hp)
        rated_kva = rated_kw / (efficiency * rated_pf)
        return rated_kva * start_mult

    @staticmethod
    def rated_current(kva: float, voltage_ll_v: float) -> float:
        """I_rated = kVA × 1000 / (√3 × V_LL)  [A]"""
        return kva * 1000.0 / (SQRT3 * voltage_ll_v)


# ─────────────────────────────────────────────────────────────────────────────
# Data Classes
# ─────────────────────────────────────────────────────────────────────────────
@dataclass
class GeneratorParams:
    """Machine parameters read from datasheet."""
    gen_id:   str   = ''
    kva:      float = 0.0
    voltage:  float = 0.0          # L-L, volts
    rated_pf: float = 0.80
    freq:     float = 60.0
    # Direct axis reactances (pu, saturated recommended for load dip)
    Xd:   float = 0.0
    Xdp:  float = 0.0             # X'd  transient
    Xdpp: float = 0.0             # X"d  subtransient
    # Quadrature axis reactances (pu)
    Xq:   float = 0.0
    Xqp:  float = 0.0             # X'q
    Xqpp: float = 0.0             # X"q
    # Resistances (pu)
    R1:   float = 0.0             # positive sequence / stator
    # Saturation coefficients (from datasheet OCC — optional, improves pre-loaded accuracy)
    # SE(E) = A_sat * exp(B_sat * E)  fitted from SE_max and SE_75max
    SE_max:   float = 0.0    # saturation factor at 1.0 pu voltage
    SE_75max: float = 0.0    # saturation factor at 0.75 pu voltage
    use_sat:  bool  = False   # toggle saturation correction on/off
    # Sequence reactances (informational — not used in balanced 3Φ dip)
    X2:   float = 0.0
    X0:   float = 0.0

    @property
    def rated_amps(self) -> float:
        return Units.rated_current(self.kva, self.voltage) if self.kva and self.voltage else 0.0


@dataclass
class StepLoadParams:
    """Inputs for a sudden step load study (no pre-existing load)."""
    load_kw:  float = 0.0
    load_pf:  float = 0.80

    @property
    def load_kva(self) -> float:
        return self.load_kw / self.load_pf if self.load_pf else 0.0

    @property
    def phi(self) -> float:
        return math.acos(self.load_pf)


@dataclass
class MotorStartParams:
    """Inputs for motor starting study (base load + motor inrush)."""
    # Pre-existing base load
    base_kva: float = 0.0
    base_pf:  float = 0.85
    # Motor starting
    motor_kva:   float = 0.0
    motor_pf:    float = 0.15
    motor_hp:    float = 0.0      # informational, used to derive kVA if needed
    motor_fla:   float = 0.0      # A at motor bus
    motor_v:     float = 0.0      # motor bus V L-L
    inrush_mult: float = 0.0      # locked rotor multiplier

    @property
    def base_phi(self) -> float:
        return math.acos(self.base_pf) if self.base_kva > 0 else 0.0

    def combined_load(self, gen_kva: float):
        """Return (S_total_pu, phi_total_rad) for the combined load phasor."""
        P = self.base_kva * self.base_pf + self.motor_kva * self.motor_pf
        Q = (self.base_kva * math.sqrt(max(0, 1 - self.base_pf  ** 2)) +
             self.motor_kva * math.sqrt(max(0, 1 - self.motor_pf ** 2)))
        S = math.sqrt(P ** 2 + Q ** 2)
        phi = math.atan2(Q, P)
        return S / gen_kva, phi


@dataclass
class PeriodResult:
    """Result for one time period (subtransient / transient / steady-state)."""
    label:    str
    Xd:       float
    Xq:       float
    Eq:       float        # pre-event internal EMF
    delta_pre: float       # pre-event load angle (rad)
    Vt:       Optional[float]  # None = unstable / not converged
    delta:    Optional[float]  # post-event load angle (rad)
    iters:    int = 0
    converged: bool = True

    @property
    def dip_pct(self) -> Optional[float]:
        return (1.0 - self.Vt) * 100.0 if self.Vt is not None else None

    @property
    def Vll(self) -> Optional[float]:
        return None   # filled in externally with gen.voltage



@dataclass
class AVRParams:
    """
    IEEE Type I (simplified DC1A) exciter / AVR parameters.

    From the CGCM configuration (Allen-Bradley or equivalent digital AVR):
        KA    — voltage regulator gain           (typ 200–400)
        TA    — regulator time constant (s)      (typ 0.01–0.05 s)
        VRMAX — ceiling (max field) voltage (pu) (typ 4–7 pu)
        VRMIN — minimum field voltage (pu)       (typ −1 to 0 pu)

    From the generator datasheet:
        KE    — exciter self-excitation constant (on datasheet)
        TE    — exciter time constant (s)        (on datasheet)

    State equations (Euler integration):
        dVR/dt = (KA × (Vref − Vt) − VR) / TA
        dEq/dt = (VR − KE × Eq) / TE

    Vref is calculated from pre-event steady-state so initial dVR/dt = 0.
    """
    KA:    float = 200.0    # voltage regulator gain
    TA:    float = 0.02     # regulator time constant (s)
    KE:    float = 1.0      # exciter self-excitation constant
    TE:    float = 0.177    # exciter time constant (s)
    VRMAX: float = 5.0      # ceiling voltage (pu)
    VRMIN: float = -1.0     # minimum field voltage (pu)
    enabled: bool = False   # toggle AVR simulation on/off


@dataclass
class StudyResult:
    """Full result set for one study."""
    study_type: str          # 'step' or 'motor'
    gen:        GeneratorParams = field(default_factory=GeneratorParams)
    periods:    list[PeriodResult] = field(default_factory=list)
    # Step load extras
    load_kva:   float = 0.0
    load_pf:    float = 0.80
    load_pu:    float = 0.0
    # Motor starting extras
    base_kva:   float = 0.0
    base_pf:    float = 0.85
    motor_kva:  float = 0.0
    motor_pf:   float = 0.15
    S_total_kva: float = 0.0
    eff_pf:     float = 0.0
    S_total_pct: float = 0.0
    avr_times:   list = field(default_factory=list)   # time vector (s)
    avr_volts:   list = field(default_factory=list)   # Vt(t) from AVR sim


# ─────────────────────────────────────────────────────────────────────────────
# Salient-Pole Newton-Raphson Solver
# ─────────────────────────────────────────────────────────────────────────────
class SalientPoleNR:
    """
    Solves salient-pole generator voltage equations using Newton-Raphson.

    Machine equations (q-axis reference frame, lagging load = positive):
        f1: E_q - V_t·cos(δ) - Ra·Iq - Xd·Id = 0
        f2: V_t·sin(δ) - Xq·Iq + Ra·Id = 0

    where:
        I  = S_pu / V_t         (constant kVA: current rises as voltage drops)
        Id = I · sin(δ + φ)     (d-axis demagnetising current)
        Iq = I · cos(δ + φ)     (q-axis current)
    """

    @staticmethod
    def sat_coeffs(SE_max: float, SE_75max: float) -> tuple[float, float]:
        """Fit exponential saturation model SE(E) = A*exp(B*E) from two datasheet points."""
        if SE_75max <= 0 or SE_max <= SE_75max:
            return 0.0, 0.0
        B = math.log(SE_max / SE_75max) / 0.25
        A = SE_75max / math.exp(0.75 * B)
        return A, B

    @staticmethod
    def SE_val(E: float, A: float, B: float) -> float:
        """Saturation factor at flux level E."""
        return A * math.exp(B * E) if A > 0 else 0.0

    @staticmethod
    def calc_Eq_pre(V_t_pre: float,
                    I_base: float,
                    phi_base: float,
                    Xd: float, Xq: float, Ra: float,
                    max_iter: int = 100) -> tuple[float, float]:
        """
        Calculate pre-event internal EMF E_q and load angle δ.

        Parameters
        ----------
        V_t_pre  : Pre-event terminal voltage (pu), typically 1.0
        I_base   : Pre-event load current magnitude (pu)
        phi_base : Pre-event load PF angle (rad), lagging positive
        Xd, Xq  : Machine reactances (pu)
        Ra       : Stator resistance (pu)

        Returns
        -------
        Eq, delta (both in pu and rad)
        """
        if I_base < 1e-9:
            return V_t_pre, 0.0

        # Closed-form starting guess (Ra = 0 approximation)
        delta = math.atan2(
            Xq * I_base * math.cos(phi_base),
            V_t_pre + Xq * I_base * math.sin(phi_base)
        )

        for _ in range(max_iter):
            Id = I_base * math.sin(delta + phi_base)
            Iq = I_base * math.cos(delta + phi_base)
            f2 = V_t_pre * math.sin(delta) - Xq * Iq + Ra * Id
            df2 = (V_t_pre * math.cos(delta)
                   - Xq * (-I_base * math.sin(delta + phi_base))
                   + Ra * (I_base * math.cos(delta + phi_base)))
            if abs(df2) < 1e-15:
                break
            step = f2 / df2
            delta -= step
            if abs(step) < 1e-12:
                break

        Id = I_base * math.sin(delta + phi_base)
        Iq = I_base * math.cos(delta + phi_base)
        Eq = V_t_pre * math.cos(delta) + Ra * Iq + Xd * Id
        return Eq, delta

    @staticmethod
    def calc_Eq_pre_sat(V_t_pre: float,
                        I_base: float,
                        phi_base: float,
                        Xd: float, Xq: float, Ra: float,
                        A_sat: float, B_sat: float) -> tuple[float, float]:
        """
        E_q with saturation correction for pre-existing load case.

        When a pre-existing base load is present, the d-axis current Id_pre
        drives the machine into saturation. The saturation correction raises
        the effective trapped flux (E_q) slightly above the linear model.

        Correction: E_q_sat = E_q_linear + SE(E_q_linear) * Xd * Id_pre

        Note: At no-load (Id_pre = 0), correction is exactly zero — correct
        behaviour since the terminal voltage = E_q = 1.0 pu at no-load.

        Accuracy: Closes ~10–20% of residual gap vs manufacturer values.
        Full dynamic saturation (remaining gap) requires the complete OCC
        curve and time-domain simulation.
        """
        Eq_lin, delta = SalientPoleNR.calc_Eq_pre(
            V_t_pre, I_base, phi_base, Xd, Xq, Ra
        )
        if A_sat <= 0 or I_base < 1e-9:
            return Eq_lin, delta
        Id_pre = I_base * math.sin(delta + phi_base)
        se_correction = SalientPoleNR.SE_val(Eq_lin, A_sat, B_sat) * Xd * Id_pre
        return Eq_lin + se_correction, delta

    @staticmethod
    def solve(Eq: float,
              S_pu: float,
              phi: float,
              Xd: float, Xq: float, Ra: float,
              V0: float = 0.90,
              delta0: float = None,
              max_iter: int = 200,
              tol: float = 1e-9) -> tuple[Optional[float], Optional[float], int, bool]:
        """
        Solve for post-disturbance terminal voltage V_t and load angle δ.

        Returns
        -------
        V_t, delta, iterations, converged
        V_t is None if the solution is unstable / did not converge.
        """
        V_t   = V0
        delta = delta0 if delta0 is not None else math.radians(15.0)

        for itr in range(max_iter):
            I  = S_pu / V_t
            Id = I * math.sin(delta + phi)
            Iq = I * math.cos(delta + phi)

            f1 = Eq - V_t * math.cos(delta) - Ra * Iq - Xd * Id
            f2 = V_t * math.sin(delta) - Xq * Iq + Ra * Id

            if abs(f1) < tol and abs(f2) < tol:
                return V_t, delta, itr, True

            dI_dV  = -I / V_t
            dId_dV = dI_dV * math.sin(delta + phi)
            dIq_dV = dI_dV * math.cos(delta + phi)
            dId_dd = I * math.cos(delta + phi)
            dIq_dd = -I * math.sin(delta + phi)

            J00 = -math.cos(delta) - Ra * dIq_dV - Xd * dId_dV
            J01 =  V_t * math.sin(delta) - Ra * dIq_dd - Xd * dId_dd
            J10 =  math.sin(delta) - Xq * dIq_dV + Ra * dId_dV
            J11 =  V_t * math.cos(delta) - Xq * dIq_dd + Ra * dId_dd

            det = J00 * J11 - J01 * J10
            if abs(det) < 1e-15:
                return None, None, itr, False

            dV = -(J11 * f1 - J01 * f2) / det
            dd = -(-J10 * f1 + J00 * f2) / det

            V_t   = max(0.005, V_t + dV)
            delta = delta + dd

        return None, None, max_iter, False


# ─────────────────────────────────────────────────────────────────────────────
# Simplified Cylindrical-Rotor Formula (for reference / quick check)
# ─────────────────────────────────────────────────────────────────────────────
def simplified_dip(X: float, I_pu: float, pf: float) -> Optional[float]:
    """
    |V_t| = √(1 − (X·I·cosφ)²) − X·I·sinφ
    Returns terminal voltage (pu), or None if unstable.
    Note: assumes cylindrical rotor (Xd only), Ra=0, E0=1.0 pu (no pre-load).
    """
    sin_phi = math.sqrt(max(0.0, 1.0 - pf ** 2))
    disc = 1.0 - (X * I_pu * pf) ** 2
    if disc < 0:
        return None
    return math.sqrt(disc) - X * I_pu * sin_phi


# ─────────────────────────────────────────────────────────────────────────────
# Study Runners
# ─────────────────────────────────────────────────────────────────────────────

def simulate_avr_response(gen: GeneratorParams,
                           avr: AVRParams,
                           S_pu: float,
                           phi: float,
                           I_base_pu: float = 0.0,
                           phi_base: float = 0.0,
                           dt: float = 0.002,
                           T_total: float = 2.5) -> tuple[list, list]:
    """
    Simulate terminal voltage recovery with IEEE Type I AVR model.

    Two-phase approach:
      Phase 1  (0 – T_sub = 50 ms)  : Subtransient window — Xd"/Xq" fixed,
                                       AVR has not yet responded.
      Phase 2  (T_sub – T_total)    : Transient+recovery — Xd'/Xq' used,
                                       AVR actively regulates terminal voltage.

    The transition at 50 ms reflects the natural decay of subtransient flux
    components, consistent with the two-period NR model.

    Returns
    -------
    times : list[float]  Time vector (s)
    volts : list[float]  Terminal voltage Vt (pu) at each time step
    """
    T_sub = 0.05

    # Pre-event E_q using Xd' (AVR operates on transient timescale)
    eq_pre_pp = SalientPoleNR.calc_Eq_pre(1.0, I_base_pu, phi_base,
                                           gen.Xdpp, gen.Xqpp, gen.R1)
    eq_pre_p  = SalientPoleNR.calc_Eq_pre(1.0, I_base_pu, phi_base,
                                           gen.Xdp,  gen.Xqp,  gen.R1)
    Eq_pp = eq_pre_pp[0]
    Eq_p  = eq_pre_p[0]

    # Initial AVR state — set so dVR/dt = 0 at t = 0
    VR0  = avr.KE * Eq_p
    Vref = 1.0 + VR0 / avr.KA

    times: list = []
    volts: list = []

    # Phase 1: subtransient window — open-loop, Eq fixed at Xd" level
    t = 0.0
    while t < T_sub - dt / 2:
        Vt = SalientPoleNR.solve(Eq_pp, S_pu, phi,
                                  gen.Xdpp, gen.Xqpp, gen.R1)[0] or 0.0
        times.append(round(t, 4))
        volts.append(Vt)
        t += dt

    # Phase 2: transient+recovery with AVR integrating
    Eq = Eq_p
    VR = VR0
    while t <= T_total + dt / 2:
        Vt = SalientPoleNR.solve(Eq, S_pu, phi,
                                  gen.Xdp, gen.Xqp, gen.R1)[0] or 0.0
        times.append(round(t, 4))
        volts.append(Vt)
        dVR = (avr.KA * (Vref - Vt) - VR) / avr.TA
        dEq  = (VR - avr.KE * Eq) / avr.TE
        VR   = max(avr.VRMIN, min(avr.VRMAX, VR + dVR * dt))
        Eq  += dEq * dt
        t   += dt

    return times, volts


def run_step_study(gen: GeneratorParams, load: StepLoadParams) -> StudyResult:
    """Step load study — generator at no-load before disturbance."""
    load_pu = load.load_kva / gen.kva
    phi     = load.phi

    result = StudyResult(
        study_type='step',
        gen=gen,
        load_kva=load.load_kva,
        load_pf=load.load_pf,
        load_pu=load_pu,
    )

    time_periods = [
        ('Subtransient  (0 – ~50 ms)',  gen.Xdpp, gen.Xqpp),
        ("Transient  (50 – ~500 ms)",   gen.Xdp,  gen.Xqp),
        ('Steady-State  (> ~1 s)',       gen.Xd,   gen.Xq),
    ]

    for label, Xd, Xq in time_periods:
        if not Xd or not Xq:
            result.periods.append(PeriodResult(
                label=label, Xd=Xd or 0, Xq=Xq or 0,
                Eq=1.0, delta_pre=0.0,
                Vt=None, delta=None, converged=False
            ))
            continue

        # Saturation correction has no effect at no-load pre-event (Id_pre=0)
        Eq, d_pre = SalientPoleNR.calc_Eq_pre(1.0, 0.0, 0.0, Xd, Xq, gen.R1)
        Vt, delta, iters, ok = SalientPoleNR.solve(Eq, load_pu, phi, Xd, Xq, gen.R1)

        result.periods.append(PeriodResult(
            label=label, Xd=Xd, Xq=Xq,
            Eq=Eq, delta_pre=d_pre,
            Vt=Vt, delta=delta, iters=iters, converged=ok
        ))

    return result


def run_motor_study(gen: GeneratorParams, motor: MotorStartParams) -> StudyResult:
    """Motor starting study — pre-existing base load + motor inrush."""
    S_pu, phi_total = motor.combined_load(gen.kva)
    S_total_kva = S_pu * gen.kva
    P_t = motor.base_kva * motor.base_pf + motor.motor_kva * motor.motor_pf
    eff_pf = P_t / S_total_kva if S_total_kva > 0 else 0.0

    I_base_pu = motor.base_kva / gen.kva if motor.base_kva > 0 else 0.0
    phi_base  = motor.base_phi

    result = StudyResult(
        study_type='motor',
        gen=gen,
        base_kva=motor.base_kva,
        base_pf=motor.base_pf,
        motor_kva=motor.motor_kva,
        motor_pf=motor.motor_pf,
        S_total_kva=S_total_kva,
        eff_pf=eff_pf,
        S_total_pct=S_total_kva / gen.kva * 100.0,
    )

    time_periods = [
        ('Subtransient  (0 – ~50 ms)',  gen.Xdpp, gen.Xqpp),
        ("Transient  (50 – ~500 ms)",   gen.Xdp,  gen.Xqp),
        ('Steady-State  (> ~1 s)',       gen.Xd,   gen.Xq),
    ]

    for label, Xd, Xq in time_periods:
        if not Xd or not Xq:
            result.periods.append(PeriodResult(
                label=label, Xd=Xd or 0, Xq=Xq or 0,
                Eq=1.0, delta_pre=0.0,
                Vt=None, delta=None, converged=False
            ))
            continue

        if gen.use_sat and gen.SE_max > 0 and gen.SE_75max > 0:
            A_s, B_s = SalientPoleNR.sat_coeffs(gen.SE_max, gen.SE_75max)
            Eq, d_pre = SalientPoleNR.calc_Eq_pre_sat(
                1.0, I_base_pu, phi_base, Xd, Xq, gen.R1, A_s, B_s
            )
        else:
            Eq, d_pre = SalientPoleNR.calc_Eq_pre(
                1.0, I_base_pu, phi_base, Xd, Xq, gen.R1
            )
        Vt, delta, iters, ok = SalientPoleNR.solve(
            Eq, S_pu, phi_total, Xd, Xq, gen.R1
        )

        result.periods.append(PeriodResult(
            label=label, Xd=Xd, Xq=Xq,
            Eq=Eq, delta_pre=d_pre,
            Vt=Vt, delta=delta, iters=iters, converged=ok
        ))

    return result


def run_motor_study_avr(gen: GeneratorParams,
                        motor: MotorStartParams,
                        avr: AVRParams) -> StudyResult:
    """Run motor starting study and attach AVR simulation times/volts."""
    result = run_motor_study(gen, motor)
    if avr.enabled and gen.Xdpp and gen.Xqpp:
        S_pu, phi = motor.combined_load(gen.kva)
        I_base_pu = motor.base_kva / gen.kva if motor.base_kva > 0 else 0.0
        phi_base  = motor.base_phi
        result.avr_times, result.avr_volts = simulate_avr_response(
            gen, avr, S_pu, phi, I_base_pu, phi_base
        )
    return result


def run_step_study_avr(gen: GeneratorParams,
                       load: StepLoadParams,
                       avr: AVRParams) -> StudyResult:
    """Run step load study and attach AVR simulation times/volts."""
    result = run_step_study(gen, load)
    if avr.enabled and gen.Xdpp and gen.Xqpp:
        result.avr_times, result.avr_volts = simulate_avr_response(
            gen, avr, result.load_pu, load.phi
        )
    return result


# ─────────────────────────────────────────────────────────────────────────────
# Console Output
# ─────────────────────────────────────────────────────────────────────────────
def print_study(result: StudyResult) -> None:
    g = result.gen
    print()
    print('=' * 78)
    if result.study_type == 'step':
        print('  STEP LOAD STUDY  —  Salient-Pole Newton-Raphson')
    else:
        print('  MOTOR STARTING STUDY  —  Salient-Pole Newton-Raphson')
    print('=' * 78)
    print(f"  Generator : {g.gen_id or '(no ID)'}")
    print(f"  Rating    : {g.kva} kVA  |  {g.voltage} V L-L  |  PF {g.rated_pf}")
    print(f"  Xd={g.Xd} pu  Xd'={g.Xdp} pu  Xd\"={g.Xdpp} pu")
    print(f"  Xq={g.Xq} pu  Xq'={g.Xqp} pu  Xq\"={g.Xqpp} pu  Ra={g.R1:.4f} pu")
    if g.use_sat and g.SE_max > 0:
        A_s, B_s = SalientPoleNR.sat_coeffs(g.SE_max, g.SE_75max)
        print(f"  Saturation: SE_max={g.SE_max}  SE_75max={g.SE_75max}  "
              f"→ SE(E) = {A_s:.5f}·exp({B_s:.3f}·E)  [ACTIVE]")
    else:
        print(f"  Saturation: disabled  (add --saturation --se-max <val> --se-75max <val> to enable)")
    print(LINE)

    if result.study_type == 'step':
        print(f"  Load: {result.load_kw if hasattr(result,'load_kw') else ''} kW  "
              f"PF {result.load_pf}  →  {result.load_kva:.1f} kVA  ({result.load_pu:.4f} pu)")
    else:
        print(f"  Base load  : {result.base_kva:.0f} kVA  PF {result.base_pf:.3f}")
        print(f"  Motor start: {result.motor_kva:.0f} kVA  PF {result.motor_pf:.3f}")
        print(f"  Combined   : {result.S_total_kva:.1f} kVA  eff PF {result.eff_pf:.3f}  "
              f"({result.S_total_pct:.1f}% of generator)")
    print(LINE)

    hdr = (f"  {'Time Period':<30} {'Xd':>6} {'Xq':>6}  "
           f"{'E₀ (pu)':>9} {'V_t (pu)':>10} {'Dip (%)':>9} {'V_LL (V)':>10}")
    print(hdr)
    print(LINE)

    for p in result.periods:
        if p.Vt and p.Vt > 0:
            dip = (1.0 - p.Vt) * 100.0
            Vll = p.Vt * g.voltage
            row = (f"  {p.label:<30} {p.Xd:>6.3f} {p.Xq:>6.3f}  "
                   f"{p.Eq:>9.5f} {p.Vt:>10.4f} {dip:>9.2f} {Vll:>10.0f}")
        else:
            row = (f"  {p.label:<30} {p.Xd:>6.3f} {p.Xq:>6.3f}  "
                   f"{p.Eq:>9.5f}   UNSTABLE / not converged")
        print(row)

    print(LINE)
    print("  Also showing simplified cylindrical-rotor formula (Ra=0, E₀=1.0) for reference:")
    for p in result.periods:
        I_pu = result.load_pu if result.study_type == 'step' else result.S_total_kva / g.kva
        pf   = result.load_pf if result.study_type == 'step' else result.eff_pf
        Vs = simplified_dip(p.Xd, I_pu, pf)
        if Vs and Vs > 0:
            print(f"    {p.label:<30} simplified V_t = {Vs:.4f} pu  ({(1-Vs)*100:.2f}% dip)")
        else:
            print(f"    {p.label:<30} simplified → UNSTABLE")
    print(LINE)
    print("  NOTE: ~1–2% remaining gap vs manufacturer is dynamic saturation (not modeled).")
    print("  Apply 10–15% margin on acceptable dip threshold for final decisions.")


# ─────────────────────────────────────────────────────────────────────────────
# Excel Report Export
# ─────────────────────────────────────────────────────────────────────────────
def export_excel(results: list[StudyResult], path: str) -> None:
    """
    Export all study results to a formatted Excel workbook.
    Each study gets its own sheet. A summary sheet is added at the front.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("ERROR: openpyxl not installed.  Run: pip install openpyxl")
        return

    wb = Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    # ── Palette ──────────────────────────────────────────────────────────────
    C = dict(
        dark_blue="1F4E79", med_blue="2E75B6", lite_blue="D6E4F7",
        pale_blue="EBF3FB", input_bg="DEEAF1", formula_bg="F2F2F2",
        pale_gray="F9F9F9", teal_bg="D4EDDA", yellow_bg="FFF3CD",
        salmon_bg="FDDBD7", white="FFFFFF", label_fg="222222",
        pass_bg="C6EFCE", fail_bg="FFC7CE",
        pass_fg="006100", fail_fg="9C0006",
    )

    def side(s='thin', c='AAAAAA'):
        return Side(style=s, color=c)

    def bdr():
        s = side()
        return Border(top=s, bottom=s, left=s, right=s)

    def hdr_fill(color=None):
        return PatternFill('solid', start_color=color or C['dark_blue'])

    def cell_style(c, bold=False, color='000000', bg=None, align='left',
                   size=10, italic=False, wrap=False, border=False, fmt=None):
        c.font = Font(name='Arial', bold=bold, italic=italic, size=size, color=color)
        if bg:
            c.fill = PatternFill('solid', start_color=bg)
        c.alignment = Alignment(horizontal=align, vertical='center',
                                indent=1 if align == 'left' else 0,
                                wrap_text=wrap)
        if border:
            c.border = bdr()
        if fmt:
            c.number_format = fmt

    # ── Summary Sheet ─────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet('Summary')
    ws_sum.column_dimensions['A'].width = 36
    for col in 'BCDE':
        ws_sum.column_dimensions[col].width = 16
    ws_sum.column_dimensions['F'].width = 28

    # Title
    ws_sum.merge_cells('A1:F1')
    c = ws_sum.cell(row=1, column=1,
        value='Generator Voltage Drop Study  —  Results Summary')
    cell_style(c, bold=True, size=14, color='FFFFFF', bg=C['dark_blue'],
               align='center')
    ws_sum.row_dimensions[1].height = 30

    # Summary table header
    for col, hdr in enumerate(
        ['Study', 'Time Period', 'Xd (pu)', 'V_terminal (pu)', 'Dip (%)', 'V_LL (V)'], 1
    ):
        c = ws_sum.cell(row=3, column=col, value=hdr)
        cell_style(c, bold=True, color='FFFFFF', bg=C['dark_blue'],
                   align='center', border=True)
    ws_sum.row_dimensions[3].height = 22

    row = 4
    period_fills = [C['teal_bg'], C['yellow_bg'], C['salmon_bg']]

    for r_idx, result in enumerate(results):
        g = result.gen
        study_label = ('Step Load' if result.study_type == 'step'
                       else 'Motor Starting')
        for p_idx, p in enumerate(result.periods):
            bg = period_fills[p_idx % 3]
            ws_sum.cell(row=row, column=1, value=study_label).fill = PatternFill('solid', start_color=bg)
            ws_sum.cell(row=row, column=1).font = Font(name='Arial', size=10, bold=True)
            ws_sum.cell(row=row, column=1).alignment = Alignment(vertical='center', indent=1)

            ws_sum.cell(row=row, column=2, value=p.label.strip()).fill = PatternFill('solid', start_color=bg)
            ws_sum.cell(row=row, column=2).font = Font(name='Arial', size=10)
            ws_sum.cell(row=row, column=2).alignment = Alignment(vertical='center', indent=1)

            for col, val in enumerate([p.Xd, p.Vt, p.dip_pct,
                                        p.Vt * g.voltage if p.Vt else None], 3):
                c = ws_sum.cell(row=row, column=col,
                                value=round(val, 4) if val is not None else 'UNSTABLE')
                c.font = Font(name='Arial', size=10, bold=(col in [4, 5]))
                c.fill = PatternFill('solid', start_color=bg)
                c.alignment = Alignment(horizontal='center', vertical='center')
                if col == 5 and p.dip_pct is not None:
                    c.number_format = '0.00'
                if col == 6 and p.Vt is not None:
                    c.number_format = '0'
            row += 1
        row += 1  # spacer between studies

    # ── Per-Study Detail Sheets ───────────────────────────────────────────────
    for result in results:
        g = result.gen
        sheet_name = ('Step Load Study' if result.study_type == 'step'
                      else 'Motor Starting Study')
        ws = wb.create_sheet(sheet_name)
        ws.column_dimensions['A'].width = 36
        for col in 'BCDEFG':
            ws.column_dimensions[col].width = 14

        # Title
        ws.merge_cells('A1:G1')
        title = ('Step Load Study' if result.study_type == 'step'
                 else 'Motor Starting Study')
        c = ws.cell(row=1, column=1,
            value=f'Generator Voltage Drop  —  {title}  (Salient-Pole NR)')
        cell_style(c, bold=True, size=13, color='FFFFFF', bg=C['dark_blue'], align='center')
        ws.row_dimensions[1].height = 28

        # Generator info block
        info_rows = [
            ('Generator ID',               g.gen_id or ''),
            ('Rating (kVA)',                g.kva),
            ('Rated Voltage (V L-L)',       g.voltage),
            ('Rated Power Factor',          g.rated_pf),
            ('Xd / Xd\' / Xd"  (pu)',      f"{g.Xd} / {g.Xdp} / {g.Xdpp}"),
            ('Xq / Xq\' / Xq"  (pu)',      f"{g.Xq} / {g.Xqp} / {g.Xqpp}"),
            ('Ra / R1  (pu)',               g.R1),
        ]
        for i, (label, val) in enumerate(info_rows):
            r = 3 + i
            lc = ws.cell(row=r, column=1, value=label)
            cell_style(lc, bold=True, bg=C['lite_blue'])
            ws.merge_cells(f'B{r}:G{r}')
            vc = ws.cell(row=r, column=2, value=val)
            cell_style(vc, bg=C['pale_blue'])
            ws.row_dimensions[r].height = 17

        # Load / motor info
        r_base = 3 + len(info_rows) + 1
        ws.merge_cells(f'A{r_base}:G{r_base}')
        c = ws.cell(row=r_base, column=1,
            value=('Load Conditions' if result.study_type == 'step'
                   else 'Motor Starting Conditions'))
        cell_style(c, bold=True, color='FFFFFF', bg=C['med_blue'])
        ws.row_dimensions[r_base].height = 20
        r_base += 1

        if result.study_type == 'step':
            load_rows = [
                ('Applied Load (kW)',         result.load_kva * result.load_pf if result.load_kva else ''),
                ('Load Power Factor',         result.load_pf),
                ('Applied Load (kVA)',        round(result.load_kva, 1)),
                ('Load Per Unit',             round(result.load_pu, 4)),
            ]
        else:
            load_rows = [
                ('Pre-existing Base Load (kVA)', result.base_kva),
                ('Base Load Power Factor',       result.base_pf),
                ('Motor Starting kVA',           result.motor_kva),
                ('Motor Starting Power Factor',  result.motor_pf),
                ('Combined Total (kVA)',          round(result.S_total_kva, 1)),
                ('Effective PF (combined)',       round(result.eff_pf, 3)),
                ('Combined as % of Generator',   round(result.S_total_pct, 1)),
            ]

        for i, (label, val) in enumerate(load_rows):
            r = r_base + i
            lc = ws.cell(row=r, column=1, value=label)
            cell_style(lc, bg=C['pale_gray'])
            ws.merge_cells(f'B{r}:G{r}')
            vc = ws.cell(row=r, column=2, value=val)
            cell_style(vc, bg=C['formula_bg'])
            ws.row_dimensions[r].height = 17

        # Results table
        r_res = r_base + len(load_rows) + 2
        for col, hdr in enumerate(
            ['Time Period', 'Xd (pu)', 'Xq (pu)', 'E₀ (pu)',
             'V_terminal (pu)', 'Voltage Dip (%)', 'V_LL (V)'], 1
        ):
            c = ws.cell(row=r_res, column=col, value=hdr)
            cell_style(c, bold=True, color='FFFFFF', bg=C['dark_blue'],
                       align='center', border=True)
        ws.row_dimensions[r_res].height = 22
        r_res += 1

        period_fills_s = [C['teal_bg'], C['yellow_bg'], C['salmon_bg']]
        for p_idx, p in enumerate(result.periods):
            bg = period_fills_s[p_idx % 3]
            vals = [
                p.label.strip(),
                round(p.Xd, 3),
                round(p.Xq, 3),
                round(p.Eq, 5),
                round(p.Vt, 4) if p.Vt else 'UNSTABLE',
                round(p.dip_pct, 2) if p.dip_pct is not None else '—',
                round(p.Vt * g.voltage, 0) if p.Vt else '—',
            ]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=r_res, column=col, value=val)
                c.font = Font(name='Arial', size=10,
                              bold=(col in [5, 6]),
                              color=(C['fail_fg'] if (col == 6 and isinstance(val, float) and val > 20)
                                     else '000000'))
                c.fill = PatternFill('solid', start_color=bg)
                c.alignment = Alignment(
                    horizontal='left' if col == 1 else 'center',
                    vertical='center', indent=1 if col == 1 else 0
                )
                c.border = bdr()
                if col == 6 and isinstance(val, float):
                    c.number_format = '0.00'
                if col == 7 and isinstance(val, float):
                    c.number_format = '0'
            ws.row_dimensions[r_res].height = 22
            r_res += 1

        # Simplified formula comparison
        r_simp = r_res + 1
        ws.merge_cells(f'A{r_simp}:G{r_simp}')
        c = ws.cell(row=r_simp, column=1,
            value='Simplified Formula Reference  (cylindrical rotor, Ra=0, E₀=1.0 pu) — shown for comparison only')
        cell_style(c, italic=True, size=9, bg='F0F0FF', color='333333')
        ws.row_dimensions[r_simp].height = 16
        r_simp += 1

        for p in result.periods:
            I_pu = (result.load_pu if result.study_type == 'step'
                    else result.S_total_kva / g.kva)
            pf   = (result.load_pf if result.study_type == 'step'
                    else result.eff_pf)
            Vs = simplified_dip(p.Xd, I_pu, pf)
            simp_txt = (f"V_t = {Vs:.4f} pu  ({(1-Vs)*100:.2f}% dip)"
                        if Vs and Vs > 0 else 'UNSTABLE')
            lc = ws.cell(row=r_simp, column=1, value=p.label.strip())
            cell_style(lc, size=9, bg='F5F5FF')
            ws.merge_cells(f'B{r_simp}:G{r_simp}')
            vc = ws.cell(row=r_simp, column=2, value=simp_txt)
            cell_style(vc, size=9, bg='F5F5FF', align='center')
            ws.row_dimensions[r_simp].height = 16
            r_simp += 1

        # Notes
        r_note = r_simp + 1
        notes = [
            '▶  Salient-pole NR closes ~60–70% of gap vs manufacturer models. '
            'Remaining ~1–2% is dynamic saturation (not modeled).',
            '▶  Apply 10–15% margin on the acceptable dip limit when using NR results for final decisions.',
            '▶  "UNSTABLE" for Xd (steady-state): generator would lose synchronism without AVR response. '
            'AVR prevents this in practice.',
        ]
        for note in notes:
            ws.merge_cells(f'A{r_note}:G{r_note}')
            c = ws.cell(row=r_note, column=1, value=note)
            cell_style(c, size=9, italic=True, bg='EBF3FB', color='1F3864', wrap=True)
            ws.row_dimensions[r_note].height = 28
            r_note += 1

    # AVR voltage recovery sheet
    avr_results = [r for r in results if r.avr_times]
    if avr_results:
        ws_avr = wb.create_sheet('AVR Response')
        ws_avr.column_dimensions['A'].width = 14
        ws_avr.column_dimensions['B'].width = 14
        ws_avr.column_dimensions['C'].width = 14
        for col, hdr in enumerate(['Time (s)', 'V_terminal (pu)', 'Dip (%)'], 1):
            c = ws_avr.cell(row=1, column=col, value=hdr)
            c.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
            c.fill = PatternFill('solid', start_color='1F4E79')
            c.alignment = Alignment(horizontal='center', vertical='center')
        for r_idx, result in enumerate(avr_results):
            col_offset = r_idx * 3
            if r_idx > 0:
                for col, hdr in enumerate(['Time (s)', 'V_terminal (pu)', 'Dip (%)'],
                                           col_offset + 1):
                    c = ws_avr.cell(row=1, column=col, value=hdr)
                    c.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
                    c.fill = PatternFill('solid', start_color='1F4E79')
                    c.alignment = Alignment(horizontal='center', vertical='center')
            for i, (t, v) in enumerate(zip(result.avr_times, result.avr_volts)):
                row = i + 2
                ws_avr.cell(row=row, column=col_offset+1, value=round(t, 4))
                ws_avr.cell(row=row, column=col_offset+2, value=round(v, 5))
                ws_avr.cell(row=row, column=col_offset+3, value=round((1-v)*100, 3))

    wb.save(path)
    print(f"\n  ✔  Report exported to: {path}")


# ─────────────────────────────────────────────────────────────────────────────
# Interactive Input Mode
# ─────────────────────────────────────────────────────────────────────────────
def prompt(label: str, default=None, cast=float) -> any:
    suffix = f"  [{default}]" if default is not None else ""
    while True:
        raw = input(f"  {label}{suffix}: ").strip()
        if not raw and default is not None:
            return default
        try:
            return cast(raw)
        except ValueError:
            print(f"    Invalid input — expected {cast.__name__}")


def interactive_mode() -> tuple[GeneratorParams, list[StudyResult]]:
    print()
    print('=' * 78)
    print('  Generator Voltage Drop Calculator  —  Interactive Mode')
    print('=' * 78)
    print()

    g = GeneratorParams()
    g.gen_id   = input("  Generator ID / Description: ").strip()
    g.kva      = prompt("Rated kVA")
    g.voltage  = prompt("Rated Voltage (V L-L)")
    g.rated_pf = prompt("Rated Power Factor", 0.80)
    g.freq     = prompt("Frequency (Hz)", 60.0)

    print()
    print("  --- Reactances (pu, saturated values recommended for load-step analysis) ---")
    print("  Note: if datasheet gives values in % — divide by 100 to enter as pu")
    g.Xd   = prompt("Xd   (Direct axis synchronous)")
    g.Xdp  = prompt("Xd'  (Direct axis transient)")
    g.Xdpp = prompt("Xd\"  (Direct axis subtransient)")
    g.Xq   = prompt("Xq   (Quadrature axis synchronous)")
    g.Xqp  = prompt("Xq'  (Quadrature axis transient)")
    g.Xqpp = prompt("Xq\"  (Quadrature axis subtransient)")
    g.R1   = prompt("R1   (Positive sequence resistance / stator resistance)", 0.0)
    g.X2   = prompt("X2   (Negative sequence, informational)", 0.0)
    g.X0   = prompt("X0   (Zero sequence, informational)", 0.0)

    results = []

    print()
    print("  --- Study type ---")
    print("  1 = Step load study only")
    print("  2 = Motor starting study only")
    print("  3 = Both")
    choice = prompt("Select", 3, int)

    if choice in (1, 3):
        print()
        print("  --- Step Load Inputs ---")
        load = StepLoadParams()
        load.load_kw = prompt("Applied load (kW)")
        load.load_pf = prompt("Load power factor", 0.80)
        results.append(run_step_study_avr(g, load, avr))

    if choice in (2, 3):
        print()
        print("  --- Motor Starting Inputs ---")
        motor = MotorStartParams()
        motor.base_kva = prompt("Pre-existing base load (kVA)  [0 = none]", 0.0)
        if motor.base_kva > 0:
            motor.base_pf = prompt("Base load power factor", 0.85)

        print()
        print("  Motor starting kVA — enter one of:")
        print("    a) Directly in kVA")
        print("    b) From FLA + inrush multiplier + motor voltage")
        print("    c) From HP + typical motor data")
        src = prompt("Choice (a/b/c)", 'a', str)

        if src.lower() == 'b':
            motor.motor_fla  = prompt("Motor FLA (A at motor bus voltage)")
            motor.inrush_mult = prompt("Inrush multiplier (e.g. 6.5 for Code F)")
            motor.motor_v    = prompt("Motor bus voltage V L-L")
            motor.motor_kva  = Units.motor_starting_kva_from_fla(
                motor.motor_fla, motor.inrush_mult, motor.motor_v
            )
            print(f"    → Starting kVA = {motor.motor_kva:.1f} kVA")
        elif src.lower() == 'c':
            motor.motor_hp = prompt("Motor HP")
            eff  = prompt("Motor efficiency", 0.95)
            rpf  = prompt("Motor rated power factor", 0.85)
            mult = prompt("Starting kVA multiplier (S_start / S_rated)", 6.0)
            motor.motor_kva = Units.motor_starting_kva_from_hp(
                motor.motor_hp, eff, rpf, mult
            )
            print(f"    → Estimated starting kVA = {motor.motor_kva:.1f} kVA")
        else:
            motor.motor_kva = prompt("Motor starting kVA")

        motor.motor_pf = prompt("Motor locked-rotor power factor", 0.15)
        results.append(run_motor_study_avr(g, motor, avr))

    return g, results


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────
def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description='Generator Voltage Drop Calculator — Salient-Pole Newton-Raphson',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )

    # Mode
    p.add_argument('--mode', choices=['step', 'motor', 'both', 'interactive'],
                   default='interactive',
                   help='Study mode (default: interactive)')

    # Generator
    p.add_argument('--gen-id',   default='', help='Generator identifier / description')
    p.add_argument('--gen-kva',  type=float, help='Generator rated kVA')
    p.add_argument('--gen-v',    type=float, help='Generator rated voltage (V L-L)')
    p.add_argument('--gen-pf',   type=float, default=0.80, help='Generator rated PF')
    p.add_argument('--freq',     type=float, default=60.0, help='Frequency (Hz)')

    # Reactances
    p.add_argument('--xd',   type=float, help='Xd  synchronous (pu)')
    p.add_argument('--xdp',  type=float, help="Xd' transient (pu)")
    p.add_argument('--xdpp', type=float, help='Xd" subtransient (pu)')
    p.add_argument('--xq',   type=float, help='Xq  synchronous (pu)')
    p.add_argument('--xqp',  type=float, help="Xq' transient (pu)")
    p.add_argument('--xqpp', type=float, help='Xq" subtransient (pu)')
    p.add_argument('--r1',   type=float, default=0.0,
                   help='R1 positive-sequence / stator resistance (pu)')
    p.add_argument('--x2',   type=float, default=0.0, help='X2 negative sequence (pu)')

    # AVR / CGCM parameters
    p.add_argument('--avr', action='store_true',
                   help='Enable IEEE Type I AVR simulation (Allen-Bradley CGCM model)')
    p.add_argument('--ka',    type=float, default=200.0,
                   help='AVR voltage regulator gain KA (typ 200–400, default 200)')
    p.add_argument('--ta',    type=float, default=0.02,
                   help='AVR regulator time constant TA (s, typ 0.01–0.05, default 0.02)')
    p.add_argument('--ke',    type=float, default=1.0,
                   help='Exciter self-excitation constant KE (from datasheet, default 1.0)')
    p.add_argument('--te',    type=float, default=0.177,
                   help='Exciter time constant TE (s, from datasheet, default 0.177)')
    p.add_argument('--vrmax', type=float, default=5.0,
                   help='Ceiling (max field) voltage VRMAX in pu (typ 4–7, default 5.0)')
    p.add_argument('--vrmin', type=float, default=-1.0,
                   help='Minimum field voltage VRMIN in pu (default −1.0)')
    # Saturation coefficients (optional)
    p.add_argument('--se-max',    type=float, default=0.0,
                   help='Saturation factor SE_max at 1.0 pu voltage (from datasheet OCC)')
    p.add_argument('--se-75max',  type=float, default=0.0,
                   help='Saturation factor SE_75max at 0.75 pu voltage (from datasheet OCC)')
    p.add_argument('--saturation', action='store_true',
                   help='Enable saturation correction (requires --se-max and --se-75max)')
    p.add_argument('--x0',   type=float, default=0.0, help='X0 zero sequence (pu)')

    # Step load
    p.add_argument('--load-kw', type=float, help='Applied load (kW)')
    p.add_argument('--load-pf', type=float, default=0.80, help='Load power factor')

    # Motor starting
    p.add_argument('--base-kva',    type=float, default=0.0,
                   help='Pre-existing base load (kVA)')
    p.add_argument('--base-pf',     type=float, default=0.85,
                   help='Base load power factor')
    p.add_argument('--motor-kva',   type=float, help='Motor starting kVA')
    p.add_argument('--motor-pf',    type=float, default=0.15,
                   help='Motor locked-rotor power factor')

    # Motor starting — from FLA / HP
    p.add_argument('--hp',          type=float, help='Motor HP (for starting kVA conversion)')
    p.add_argument('--fla',         type=float, help='Motor FLA (A) at motor bus voltage')
    p.add_argument('--v-motor',     type=float, help='Motor bus voltage (V L-L)')
    p.add_argument('--inrush-mult', type=float, help='Locked-rotor inrush multiplier')
    p.add_argument('--motor-eff',   type=float, default=0.95,
                   help='Motor efficiency (for HP method, default 0.95)')
    p.add_argument('--motor-rated-pf', type=float, default=0.85,
                   help='Motor rated PF (for HP method, default 0.85)')
    p.add_argument('--start-mult',  type=float, default=6.0,
                   help='Starting kVA multiplier for HP method (default 6.0)')

    # Output
    p.add_argument('--export', metavar='FILE.xlsx',
                   help='Export results to Excel workbook')

    # Unit conversion
    p.add_argument('--convert', nargs='+', metavar='TYPE',
                   help='Unit conversion mode: hp <val> | pct <val> | '
                        'starting-kva <FLA> <mult> <V_LL> | ohms <val> <kva> <v>')

    return p


def main() -> None:
    parser = build_parser()
    args   = parser.parse_args()

    # ── Unit conversion mode ─────────────────────────────────────────────────
    if args.convert:
        ctype = args.convert[0].lower()
        try:
            if ctype == 'hp' and len(args.convert) >= 2:
                hp = float(args.convert[1])
                print(f"  {hp} HP  =  {Units.hp_to_kw(hp):.4f} kW")
            elif ctype == 'kw' and len(args.convert) >= 2:
                kw = float(args.convert[1])
                print(f"  {kw} kW  =  {Units.kw_to_hp(kw):.3f} HP")
            elif ctype == 'pct' and len(args.convert) >= 2:
                pct = float(args.convert[1])
                print(f"  {pct}%  =  {Units.pct_to_pu(pct):.5f} pu")
            elif ctype == 'starting-kva' and len(args.convert) >= 4:
                fla, mult, v = (float(args.convert[1]),
                                float(args.convert[2]),
                                float(args.convert[3]))
                kva = Units.motor_starting_kva_from_fla(fla, mult, v)
                print(f"  {fla} A  ×  {mult}  ×  {v} V  =  {kva:.1f} kVA starting")
            elif ctype == 'ohms' and len(args.convert) >= 4:
                z, kva, v = (float(args.convert[1]),
                              float(args.convert[2]),
                              float(args.convert[3]))
                print(f"  {z} Ω  =  {Units.ohms_to_pu(z, kva, v):.5f} pu  "
                      f"(Z_base = {Units.base_impedance(kva, v):.4f} Ω)")
            else:
                print("Unknown conversion. Options: hp | kw | pct | starting-kva | ohms")
        except (ValueError, IndexError) as e:
            print(f"Conversion error: {e}")
        return

    # ── Interactive mode ─────────────────────────────────────────────────────
    if args.mode == 'interactive' or not args.gen_kva:
        _, results = interactive_mode()
        for r in results:
            print_study(r)
        if args.export:
            export_excel(results, args.export)
        return

    # ── CLI mode ─────────────────────────────────────────────────────────────
    g = GeneratorParams(
        gen_id=args.gen_id,
        kva=args.gen_kva,
        voltage=args.gen_v,
        rated_pf=args.gen_pf,
        freq=args.freq,
        Xd=args.xd   or 0.0,
        Xdp=args.xdp  or 0.0,
        Xdpp=args.xdpp or 0.0,
        Xq=args.xq   or 0.0,
        Xqp=args.xqp  or 0.0,
        Xqpp=args.xqpp or 0.0,
        R1=args.r1,
        X2=args.x2,
        X0=args.x0,
        SE_max=args.se_max,
        SE_75max=args.se_75max,
        use_sat=args.saturation,
    )

    avr = AVRParams(
        KA=args.ka, TA=args.ta, KE=args.ke, TE=args.te,
        VRMAX=args.vrmax, VRMIN=args.vrmin,
        enabled=args.avr,
    )

    results = []

    # Resolve motor starting kVA from FLA or HP if direct kVA not given
    motor_kva = args.motor_kva
    if not motor_kva:
        if args.fla and args.inrush_mult and args.v_motor:
            motor_kva = Units.motor_starting_kva_from_fla(
                args.fla, args.inrush_mult, args.v_motor
            )
            print(f"\n  Motor starting kVA from FLA: "
                  f"{args.fla}A × {args.inrush_mult} × {args.v_motor}V "
                  f"= {motor_kva:.1f} kVA")
        elif args.hp:
            motor_kva = Units.motor_starting_kva_from_hp(
                args.hp, args.motor_eff, args.motor_rated_pf, args.start_mult
            )
            print(f"\n  Estimated motor starting kVA from {args.hp} HP: {motor_kva:.1f} kVA")

    if args.mode in ('step', 'both') and args.load_kw:
        load = StepLoadParams(load_kw=args.load_kw, load_pf=args.load_pf)
        results.append(run_step_study_avr(g, load, avr))

    if args.mode in ('motor', 'both') and motor_kva:
        motor = MotorStartParams(
            base_kva=args.base_kva,
            base_pf=args.base_pf,
            motor_kva=motor_kva,
            motor_pf=args.motor_pf,
        )
        results.append(run_motor_study_avr(g, motor, avr))

    if not results:
        print("\nNo studies run — check inputs. Use --help for usage.")
        sys.exit(1)

    for r in results:
        print_study(r)

    if args.export:
        export_excel(results, args.export)


if __name__ == '__main__':
    main()
